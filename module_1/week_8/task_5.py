import pandas as pd
import numpy as np
import warnings

warnings.filterwarnings('ignore')


def load_data():
    """Load dwelling and sales data."""

    print("Loading data files...")

    # Load regional dwelling data from Task 3
    dwellings_df = pd.read_csv('census_dwelling_regional_summary.csv')

    # Load regional property sales from Task 4
    sales_df = pd.read_csv('task_4_regional_property_summary.csv')

    print(f"✓ Dwelling data loaded: {len(dwellings_df)} regions")
    print(f"✓ Sales data loaded: {len(sales_df)} regions")

    return dwellings_df, sales_df


def merge_datasets(dwellings_df, sales_df):
    """Merge dwelling and sales data by region."""

    # Merge on Region_Code and Region_Name
    merged = dwellings_df.merge(
        sales_df,
        on=['Region_Code', 'Region_Name'],
        how='outer',
        suffixes=('_Dwellings', '_Sales')
    )

    # Fill NaN values with 0 for sales (regions with no sales)
    sales_cols = ['D', 'F', 'S', 'T', 'Total']
    for col in sales_cols:
        if col in merged.columns:
            merged[col] = merged[col].fillna(0).astype(int)

    print(f"\n=== Merged Regional Data ===")
    print(merged[['Region_Name', 'Total_Dwellings', 'Total']])

    return merged


def calculate_percentages(merged_df):
    """Calculate percentage of each property type sold by region and nationally."""

    # Rename sales columns for clarity
    merged_df = merged_df.rename(columns={
        'D': 'Sales_Detached',
        'F': 'Sales_Flats',
        'S': 'Sales_Semi',
        'T': 'Sales_Terraced',
        'Total': 'Total_Sales'
    })

    # Calculate percentages for each property type by region
    merged_df['Pct_Detached'] = (merged_df['Sales_Detached'] / merged_df['Total_Sales'] * 100).round(2)
    merged_df['Pct_Flats'] = (merged_df['Sales_Flats'] / merged_df['Total_Sales'] * 100).round(2)
    merged_df['Pct_Semi'] = (merged_df['Sales_Semi'] / merged_df['Total_Sales'] * 100).round(2)
    merged_df['Pct_Terraced'] = (merged_df['Sales_Terraced'] / merged_df['Total_Sales'] * 100).round(2)

    # Calculate sales as percentage of total dwellings
    merged_df['Sales_Rate'] = (merged_df['Total_Sales'] / merged_df['Total_Dwellings'] * 100).round(2)

    # Handle division by zero (regions with no sales)
    merged_df['Pct_Detached'] = merged_df['Pct_Detached'].fillna(0)
    merged_df['Pct_Flats'] = merged_df['Pct_Flats'].fillna(0)
    merged_df['Pct_Semi'] = merged_df['Pct_Semi'].fillna(0)
    merged_df['Pct_Terraced'] = merged_df['Pct_Terraced'].fillna(0)
    merged_df['Sales_Rate'] = merged_df['Sales_Rate'].fillna(0)

    return merged_df


def calculate_national_totals(merged_df):
    """Calculate national totals and percentages for England and Wales."""

    # Sum all regions to get national totals
    national = pd.DataFrame({
        'Region_Code': ['NATIONAL'],
        'Region_Name': ['England & Wales'],
        'Total_Dwellings': [merged_df['Total_Dwellings'].sum()],
        'Unshared_Dwellings': [merged_df['Unshared_Dwellings'].sum()],
        'Shared_Dwellings': [merged_df['Shared_Dwellings'].sum()],
        'Sales_Detached': [merged_df['Sales_Detached'].sum()],
        'Sales_Flats': [merged_df['Sales_Flats'].sum()],
        'Sales_Semi': [merged_df['Sales_Semi'].sum()],
        'Sales_Terraced': [merged_df['Sales_Terraced'].sum()],
        'Total_Sales': [merged_df['Total_Sales'].sum()]
    })

    # Calculate national percentages
    total_sales = national['Total_Sales'].iloc[0]
    national['Pct_Detached'] = (national['Sales_Detached'] / total_sales * 100).round(2)
    national['Pct_Flats'] = (national['Sales_Flats'] / total_sales * 100).round(2)
    national['Pct_Semi'] = (national['Sales_Semi'] / total_sales * 100).round(2)
    national['Pct_Terraced'] = (national['Sales_Terraced'] / total_sales * 100).round(2)
    national['Sales_Rate'] = (national['Total_Sales'] / national['Total_Dwellings'] * 100).round(2)

    # Combine regional and national data
    combined = pd.concat([merged_df, national], ignore_index=True)

    print("\n=== National Summary ===")
    print(f"Total Dwellings: {national['Total_Dwellings'].iloc[0]:,}")
    print(f"Total Sales: {national['Total_Sales'].iloc[0]:,}")
    print(f"National Sales Rate: {national['Sales_Rate'].iloc[0]}%")
    print(f"\nNational Property Type Distribution:")
    print(f"  Detached: {national['Pct_Detached'].iloc[0]}%")
    print(f"  Flats: {national['Pct_Flats'].iloc[0]}%")
    print(f"  Semi-detached: {national['Pct_Semi'].iloc[0]}%")
    print(f"  Terraced: {national['Pct_Terraced'].iloc[0]}%")

    return combined


def identify_maxima(df):
    """Identify and flag maximum values for each metric."""

    # Exclude national row for maxima calculation
    regional_only = df[df['Region_Code'] != 'NATIONAL'].copy()

    # Find maxima for each percentage column
    max_detached = regional_only['Pct_Detached'].max()
    max_flats = regional_only['Pct_Flats'].max()
    max_semi = regional_only['Pct_Semi'].max()
    max_terraced = regional_only['Pct_Terraced'].max()
    max_sales_rate = regional_only['Sales_Rate'].max()

    # Create flags for maxima
    df['Max_Detached'] = df['Pct_Detached'] == max_detached
    df['Max_Flats'] = df['Pct_Flats'] == max_flats
    df['Max_Semi'] = df['Pct_Semi'] == max_semi
    df['Max_Terraced'] = df['Pct_Terraced'] == max_terraced
    df['Max_Sales_Rate'] = df['Sales_Rate'] == max_sales_rate

    print("\n=== Regional Maxima ===")
    print(
        f"Highest % Detached: {regional_only.loc[regional_only['Pct_Detached'].idxmax(), 'Region_Name']} ({max_detached}%)")
    print(f"Highest % Flats: {regional_only.loc[regional_only['Pct_Flats'].idxmax(), 'Region_Name']} ({max_flats}%)")
    print(
        f"Highest % Semi-detached: {regional_only.loc[regional_only['Pct_Semi'].idxmax(), 'Region_Name']} ({max_semi}%)")
    print(
        f"Highest % Terraced: {regional_only.loc[regional_only['Pct_Terraced'].idxmax(), 'Region_Name']} ({max_terraced}%)")
    print(
        f"Highest Sales Rate: {regional_only.loc[regional_only['Sales_Rate'].idxmax(), 'Region_Name']} ({max_sales_rate}%)")

    return df


def create_summary_table(df):
    """Create a clean summary table for output."""

    # Select and order columns
    summary = df[[
        'Region_Code',
        'Region_Name',
        'Total_Dwellings',
        'Sales_Detached',
        'Sales_Flats',
        'Sales_Semi',
        'Sales_Terraced',
        'Total_Sales',
        'Pct_Detached',
        'Pct_Flats',
        'Pct_Semi',
        'Pct_Terraced',
        'Sales_Rate',
        'Max_Detached',
        'Max_Flats',
        'Max_Semi',
        'Max_Terraced',
        'Max_Sales_Rate'
    ]].copy()

    return summary


def save_to_excel_with_formatting(df, filename):
    """Save to Excel with conditional formatting highlighting maxima."""

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font

        # Save to Excel
        df.to_excel(filename, index=False, sheet_name='Regional Analysis')

        # Load workbook for formatting
        wb = load_workbook(filename)
        ws = wb['Regional Analysis']

        # Define highlight fill
        highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        bold_font = Font(bold=True)

        # Find column indices for percentage columns
        pct_cols = {
            'Pct_Detached': None,
            'Pct_Flats': None,
            'Pct_Semi': None,
            'Pct_Terraced': None,
            'Sales_Rate': None
        }

        max_cols = {
            'Max_Detached': None,
            'Max_Flats': None,
            'Max_Semi': None,
            'Max_Terraced': None,
            'Max_Sales_Rate': None
        }

        # Get column indices
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value in pct_cols:
                pct_cols[cell.value] = idx
            if cell.value in max_cols:
                max_cols[cell.value] = idx

        # Apply highlighting based on Max flags
        for row_idx in range(2, ws.max_row + 1):
            # Detached
            if ws.cell(row_idx, max_cols['Max_Detached']).value == True:
                ws.cell(row_idx, pct_cols['Pct_Detached']).fill = highlight_fill
                ws.cell(row_idx, pct_cols['Pct_Detached']).font = bold_font

            # Flats
            if ws.cell(row_idx, max_cols['Max_Flats']).value == True:
                ws.cell(row_idx, pct_cols['Pct_Flats']).fill = highlight_fill
                ws.cell(row_idx, pct_cols['Pct_Flats']).font = bold_font

            # Semi
            if ws.cell(row_idx, max_cols['Max_Semi']).value == True:
                ws.cell(row_idx, pct_cols['Pct_Semi']).fill = highlight_fill
                ws.cell(row_idx, pct_cols['Pct_Semi']).font = bold_font

            # Terraced
            if ws.cell(row_idx, max_cols['Max_Terraced']).value == True:
                ws.cell(row_idx, pct_cols['Pct_Terraced']).fill = highlight_fill
                ws.cell(row_idx, pct_cols['Pct_Terraced']).font = bold_font

            # Sales Rate
            if ws.cell(row_idx, max_cols['Max_Sales_Rate']).value == True:
                ws.cell(row_idx, pct_cols['Sales_Rate']).fill = highlight_fill
                ws.cell(row_idx, pct_cols['Sales_Rate']).font = bold_font

        # Hide the Max columns
        for col_letter in ['N', 'O', 'P', 'Q', 'R']:
            ws.column_dimensions[col_letter].hidden = True

        # Make header bold
        for cell in ws[1]:
            cell.font = bold_font

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filename)
        print(f"\n✓ Excel file with formatting saved to: {filename}")

    except ImportError:
        # Fallback: save without formatting
        df.to_excel(filename, index=False, sheet_name='Regional Analysis')
        print(f"\n✓ Excel file saved to: {filename} (without formatting - openpyxl not installed)")


def main():
    """Main function for Task 5."""

    print("=" * 70)
    print("TASK 5: Regional Property Analysis - Dwellings vs Sales")
    print("=" * 70)
    print()

    # Load data
    dwellings_df, sales_df = load_data()

    # Merge datasets
    merged_df = merge_datasets(dwellings_df, sales_df)

    # Calculate percentages
    print("\nCalculating percentages...")
    analysis_df = calculate_percentages(merged_df)

    # Add national totals
    print("\nCalculating national totals...")
    complete_df = calculate_national_totals(analysis_df)

    # Identify maxima
    print("\nIdentifying regional maxima...")
    final_df = identify_maxima(complete_df)

    # Create summary table
    summary_table = create_summary_table(final_df)

    # Display summary
    print("\n=== Regional Analysis Summary Table ===")
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    display_cols = ['Region_Name', 'Total_Sales', 'Pct_Detached', 'Pct_Flats', 'Pct_Semi', 'Pct_Terraced', 'Sales_Rate']
    print(summary_table[display_cols].to_string(index=False))

    # Save outputs
    print("\n\nSaving outputs...")
    summary_table.to_csv('regional_analysis_complete.csv', index=False)
    print("✓ CSV saved to: regional_analysis_complete.csv")

    # Save to Excel with formatting
    save_to_excel_with_formatting(summary_table, 'regional_analysis_complete.xlsx')

    print("\n" + "=" * 70)
    print("Task 5 Complete!")
    print("=" * 70)


if __name__ == "__main__":
    main()